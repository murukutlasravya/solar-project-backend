from pathlib import Path
from typing import List, Tuple

from pypdf import PdfReader
import docx
import openpyxl


def extract_pdf_text(path: Path) -> List[Tuple[int, str]]:
    """
    Extract text from a PDF file.

    Returns:
        list of (page_number, text)
    """
    reader = PdfReader(str(path))
    results: List[Tuple[int, str]] = []

    for i, page in enumerate(reader.pages):
        txt = (page.extract_text() or "").strip()
        if not txt:
            continue
        # page numbers are 1-based
        results.append((i + 1, txt))

    return results


def extract_docx_text(path: Path) -> List[Tuple[int, str]]:
    """
    Extract text from a Word (.docx) file.

    Returns:
        list of (section_index, text)

    We group paragraphs into chunks of ~10 paragraphs to avoid
    creating too many tiny chunks.
    """
    doc = docx.Document(str(path))
    chunks: List[Tuple[int, str]] = []
    buffer: List[str] = []
    section_idx = 1

    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue

        buffer.append(text)

        # Simple heuristic: every 10 paragraphs, flush a chunk
        if len(buffer) >= 10:
            chunks.append((section_idx, "\n".join(buffer)))
            section_idx += 1
            buffer = []

    # Flush remaining buffer
    if buffer:
        chunks.append((section_idx, "\n".join(buffer)))

    return chunks


def extract_xlsx_text(path: Path) -> List[Tuple[int, str]]:
    """
    Extract text-like content from an Excel workbook (.xlsx, .xlsm, .xls).

    Returns:
        list of (sheet_index, text)

    Each sheet becomes one big text blob with lines like:
      'Sheet: Sheet1'
      'Row: value1 | value2 | ...'
    """
    wb = openpyxl.load_workbook(str(path), data_only=True)
    results: List[Tuple[int, str]] = []

    for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]
        lines: List[str] = [f"Sheet: {sheet_name}"]

        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c not in (None, "")]
            if not cells:
                continue
            lines.append(" | ".join(cells))

        text = "\n".join(lines).strip()
        if text:
            results.append((sheet_idx, text))

    return results
